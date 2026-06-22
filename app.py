from __future__ import annotations

import csv
import io
import os
import sqlite3
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from flask import Flask, abort, g, jsonify, render_template, request, send_file

try:
    import psycopg
    from psycopg import IntegrityError as PsycopgIntegrityError
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover - optional for local SQLite usage
    psycopg = None
    PsycopgIntegrityError = None
    dict_row = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover - handled at runtime via requirements.txt
    Workbook = None
    Alignment = Font = PatternFill = None


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("APP_DATA_DIR", str(BASE_DIR / "data")))
DB_PATH = Path(os.getenv("DB_PATH", str(DATA_DIR / "support_tracker.db")))
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)
RUNNING_ON_RENDER = any(
    os.getenv(variable_name, "").strip()
    for variable_name in ("RENDER", "RENDER_EXTERNAL_URL", "RENDER_SERVICE_ID")
)
DB_INTEGRITY_ERRORS = (sqlite3.IntegrityError,)
if PsycopgIntegrityError is not None:
    DB_INTEGRITY_ERRORS = DB_INTEGRITY_ERRORS + (PsycopgIntegrityError,)
APP_TITLE = "Nhật ký hỗ trợ khách hàng"
CHANNEL_OPTIONS = [
    "Group ĐTP-AM-Hỗ trợ",
    "Cộng đồng hóa đơn",
    "Điện thoại",
    "Nhắn tin riêng",
    "Khác",
]
STATUS_OPTIONS = [
    "Đã xử lý",
    "Đang xử lý",
    "Chờ phản hồi",
]
SUPPORTER_OPTIONS = [
    {"name": "Hằng", "area": ""},
    {"name": "Duyên", "area": ""},
]
REQUESTER_OPTIONS = [
    {"name": "Trần Thị Ngọc Huệ", "area": "CLH"},
    {"name": "Ngô Hoài Trinh", "area": "CLH"},
    {"name": "Ngô Minh Nhã", "area": "CLH"},
    {"name": "Nguyễn Võ Trọng Nghĩa", "area": "HNU"},
    {"name": "Phạm Lê Thuận An", "area": "HNU"},
    {"name": "Thái Vệ Thành", "area": "HNU"},
    {"name": "Huỳnh Văn Dư", "area": "LVG"},
    {"name": "Nguyễn Hoàng Trung Nghĩa", "area": "LVG"},
    {"name": "Huỳnh Ngọc Trung", "area": "LVG"},
    {"name": "Lưu Khánh Hào", "area": "LVG"},
    {"name": "Phan Hồng Ân", "area": "MTH"},
    {"name": "Nguyễn Thanh Nhàn", "area": "SDC"},
    {"name": "Lê Thành Tâm", "area": "SDC"},
    {"name": "Võ Quốc Vân", "area": "SDC"},
    {"name": "Nguyễn Minh Trí", "area": "TNG"},
    {"name": "Lưu Hớn Đạt", "area": "TBH"},
    {"name": "Trần Minh Quân", "area": "TBH"},
    {"name": "Lê Hữu Hiệp", "area": "TMI"},
    {"name": "Lê Thanh Vũ", "area": "TMI"},
    {"name": "Nguyễn Thị Mỹ Ánh", "area": "MTO"},
    {"name": "Bùi Thị Mỹ Hạnh", "area": "MTO"},
    {"name": "Nguyễn Văn Rề Phiên", "area": "CBE"},
    {"name": "Nguyễn Thanh Phong", "area": "CLY"},
    {"name": "Nguyễn Thị Như Ý", "area": "CLY"},
    {"name": "Huỳnh Lạc Nguyên", "area": "CTH"},
    {"name": "Nguyễn Quốc Vương", "area": "CTH"},
    {"name": "Nguyễn Tuấn Hải", "area": "CGO"},
    {"name": "Hồ Văn Tâm", "area": "GCT"},
    {"name": "Trương Thị Anh Đào", "area": "GCG"},
    {"name": "Nguyễn Văn Tăng Tài", "area": "GCG"},
    {"name": "Nguyễn Chí Tâm", "area": "TPC"},
]
REQUESTER_OPTION_NAMES = [option["name"] for option in REQUESTER_OPTIONS]

DEFAULT_SERVICES = [
    "VNPT CA",
    "VNPT BHXH",
    "VNPT-Invoice",
    "VNPT-HKD",
    "VNPT eReceipt",
    "VNPT-Invoice Inbot",
    "VNPT-Pharmacy",
    "Hợp đồng điện tử",
    "Dịch vụ CNTT khác",
]

SERVICE_CONTENT_SUGGESTION_GROUPS = [
    {
        "aliases": ["VNPT CA"],
        "suggestions": [
            "Chữ ký số mới: đăng ký tài khoản, cập nhật seri ký số.",
            "Chữ ký số gia hạn: cập nhật chữ ký số cho kế toán, lãnh đạo.",
            "Hỗ trợ đăng ký tài khoản tiền gửi, cập nhật seri ký số.",
            "Hỗ trợ tất toán tài khoản, cập nhật seri ký số.",
            "Xử lý lỗi ký số khi duyệt chứng từ kho bạc.",
            "Hỗ trợ đăng ký/ thay đổi/ hủy Ủy quyền điện, nước, viễn thông.",
            "Hỗ trợ ký số/ cập nhật ký số trang báo cáo kho bạc https://bctcnn.vst.mof.gov.vn/.",
            "Chữ ký số mới: đăng ký tài khoản, cập nhật seri ký số trên trang thuế điện tử.",
            "Chữ ký số gia hạn: cập nhật chữ ký số trên trang thuế điện tử.",
            "Cài đặt phần mềm kê khai thuế, phần mềm đọc tờ khai thuế, plugin ký số trên trang thuế.",
            "Xử lý lỗi ký số khi nộp tờ khai.",
            "Hướng dẫn cập nhật ký số trên các trang web khác: Cổng Dịch vụ công Quốc gia, Hợp đồng điện tử, Hóa đơn điện tử, Trang đấu thầu, file pdf,...",
            "Cài đặt phần mềm ký số, xử lý các lỗi ký số khác như nhập sai mã pin bị khóa, máy tính không nhận cổng USB chữ ký số.",
        ],
    },
    {
        "aliases": ["VNPT BHXH", "VNPT-BHXH"],
        "suggestions": [
            "Cài đặt phần mềm, plugin ký số.",
            "Hướng dẫn sử dụng các chức năng: tạo mới, nhập thông tin, ký số, gửi hồ sơ, tra cứu hồ sơ, cập nhật thông tin đơn vị.",
            "Hướng dẫn nghiệp vụ trên phần mềm: báo tăng, báo giảm, điều chỉnh đóng BHXH; đăng ký, thay đổi thông tin đóng BHXH; giải quyết hưởng chế độ thai sản; giải quyết hưởng trợ cấp dưỡng sức, phục hồi sức khỏe sau ốm đau, thai sản, tai nạn lao động, bệnh nghề nghiệp.",
            "Chữ ký số mới: đăng ký tài khoản.",
            "Chữ ký số gia hạn: cập nhật chữ ký số.",
            "Xử lý các lỗi kê khai, lỗi ký số, hỗ trợ sao lưu, khôi phục dữ liệu từ máy cũ qua máy mới.",
        ],
    },
    {
        "aliases": ["VNPT-Invoice Inbot", "Invoice Inbot", "Inbot"],
        "suggestions": [
            "Hướng dẫn sử dụng các chức năng: đăng nhập, quên mật khẩu, tra cứu hóa đơn đầu vào.",
            "Hỗ trợ thay đổi thông tin, reset mật khẩu.",
            "Cấu hình đồng bộ hóa đơn đầu vào từ CQT, NCC, email.",
        ],
    },
    {
        "aliases": ["VNPT-Pharmacy", "Pharmacy"],
        "suggestions": [
            "Hướng dẫn sử dụng các chức năng cơ bản: đăng nhập, quên mật khẩu.",
        ],
    },
    {
        "aliases": [
            "VNPT-Invoice",
            "VNPT-HKD",
            "VNPT eReceipt",
            "eReceipt",
        ],
        "suggestions": [
            "Hướng dẫn sử dụng các chức năng: đăng nhập, tạo mới, phát hành, thay thế, điều chỉnh tăng/ giảm doanh thu, hủy, gửi thông báo sai sót, upload, tra cứu, tải, in hàng loạt, chuyển đổi lưu trữ, xuất báo cáo thống kê hóa đơn/ biên lai/ chứng từ khấu trừ thuế.",
            "Cài đặt plugin ký số.",
            "Xử lý lỗi ký số phát hành hóa đơn/ biên lai/ chứng từ khấu trừ thuế, lỗi không cấp mã CQT, lỗi không tiếp nhận,...",
            "Hỗ trợ chuyển dải ký hiệu năm mới, thay đổi thông tin, reset mật khẩu.",
            "Cấu hình mẫu, ký hiệu hóa đơn/ biên lai/ chứng từ khấu trừ thuế, cấu hình menu/ đơn vị tính/ hình thức thanh toán/ hiển thị theo yêu cầu khách hàng.",
            "Hỗ trợ chuyển hóa đơn có mã cơ quan thuế sang hóa đơn có mã khởi tạo từ máy tính tiền, cấu hình mẫu hóa đơn, cấu hình hệ thống, cấu hình dải ký hiệu.",
        ],
    },
]

SQLITE_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS services (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    description TEXT NOT NULL DEFAULT '',
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS support_records (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    support_date TEXT NOT NULL,
    customer_name TEXT NOT NULL,
    requester_name TEXT NOT NULL DEFAULT '',
    service_id INTEGER NOT NULL,
    support_content TEXT NOT NULL,
    channel TEXT NOT NULL,
    supporter_name TEXT NOT NULL,
    status TEXT NOT NULL,
    notes TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(service_id) REFERENCES services(id)
);

CREATE INDEX IF NOT EXISTS idx_support_date ON support_records(support_date);
CREATE INDEX IF NOT EXISTS idx_support_service ON support_records(service_id);
CREATE INDEX IF NOT EXISTS idx_support_status ON support_records(status);
CREATE INDEX IF NOT EXISTS idx_support_supporter ON support_records(supporter_name);
"""

POSTGRES_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS services (
    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
    name TEXT NOT NULL UNIQUE,
    description TEXT NOT NULL DEFAULT '',
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS support_records (
    id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
    support_date TEXT NOT NULL,
    customer_name TEXT NOT NULL,
    requester_name TEXT NOT NULL DEFAULT '',
    service_id INTEGER NOT NULL REFERENCES services(id),
    support_content TEXT NOT NULL,
    channel TEXT NOT NULL,
    supporter_name TEXT NOT NULL,
    status TEXT NOT NULL,
    notes TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_support_date ON support_records(support_date);
CREATE INDEX IF NOT EXISTS idx_support_service ON support_records(service_id);
CREATE INDEX IF NOT EXISTS idx_support_status ON support_records(status);
CREATE INDEX IF NOT EXISTS idx_support_supporter ON support_records(supporter_name);
"""


app = Flask(__name__)
app.json.ensure_ascii = False
app.config["TEMPLATES_AUTO_RELOAD"] = True


def normalize_database_url(value: str) -> str:
    if value.startswith("postgres://"):
        return f"postgresql://{value[len('postgres://'):]}"
    return value


def db_backend_name() -> str:
    return "postgres" if USE_POSTGRES else "sqlite"


def active_schema_sql() -> str:
    return POSTGRES_SCHEMA_SQL if USE_POSTGRES else SQLITE_SCHEMA_SQL


def adapt_sql(sql: str) -> str:
    if not USE_POSTGRES:
        return sql
    return sql.replace("?", "%s")


def execute_query(db: Any, sql: str, params: tuple[Any, ...] | list[Any] | None = None):
    return db.execute(adapt_sql(sql), params or ())


def execute_many(db: Any, sql: str, params_seq: list[tuple[Any, ...]]):
    if USE_POSTGRES:
        with db.cursor() as cursor:
            cursor.executemany(adapt_sql(sql), params_seq)
        return None
    return db.executemany(sql, params_seq)


def execute_script(db: Any, script: str) -> None:
    if not USE_POSTGRES:
        db.executescript(script)
        return

    for statement in script.split(";"):
        cleaned = statement.strip()
        if cleaned:
            db.execute(cleaned)


def iso_now() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def current_week_range(reference: date | None = None) -> tuple[date, date]:
    reference = reference or date.today()
    start = reference - timedelta(days=reference.weekday())
    end = start + timedelta(days=6)
    return start, end


def month_range(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    return start, next_month - timedelta(days=1)


def format_display_date(value: str | None) -> str:
    if not value:
        return ""
    return date.fromisoformat(value).strftime("%d/%m/%Y")


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_lookup_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    normalized = normalized.replace("đ", "d").replace("Đ", "D")
    without_marks = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    cleaned = "".join(
        character.lower() if character.isalnum() else " "
        for character in without_marks
    )
    return " ".join(cleaned.split())


def canonical_channel_name(value: Any) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""

    legacy_map = {
        "Ultra kh\u00e1ch": "Group \u0110TP-AM-H\u1ed7 tr\u1ee3",
        "Zalo": "Nh\u1eafn tin ri\u00eang",
        "Tr\u1ef1c ti\u1ebfp": "Kh\u00e1c",
        "C\u1ed9ng \u0111\u1ed3ng H\u00f3a \u0111\u01a1n": "C\u1ed9ng \u0111\u1ed3ng h\u00f3a \u0111\u01a1n",
        "C\u00f4ng \u0111\u1ed3ng H\u00f3a \u0111\u01a1n": "C\u1ed9ng \u0111\u1ed3ng h\u00f3a \u0111\u01a1n",
    }
    return legacy_map.get(raw, raw)


def short_text(value: str, limit: int = 90) -> str:
    cleaned = normalize_text(value)
    if len(cleaned) <= limit:
        return cleaned
    return f"{cleaned[: limit - 3]}..."


def format_counter(counter: Counter[str], labels: list[str] | None = None) -> str:
    if not counter:
        return "không phát sinh"
    order = labels or list(counter.keys())
    parts: list[str] = []
    for label in order:
        count = counter.get(label, 0)
        if count:
            parts.append(f"{label}: {count}")
    if not parts:
        parts = [f"{label}: {count}" for label, count in counter.items() if count]
    return ", ".join(parts)


def get_db():
    if "db" not in g:
        if USE_POSTGRES:
            if psycopg is None:
                raise RuntimeError(
                    "DATABASE_URL is set but psycopg is not installed. "
                    "Run pip install -r requirements.txt."
                )
            g.db = psycopg.connect(
                normalize_database_url(DATABASE_URL),
                row_factory=dict_row,
            )
        else:
            if RUNNING_ON_RENDER:
                raise RuntimeError(
                    "DATABASE_URL is not configured. "
                    "On Render, configure a PostgreSQL database to avoid losing data."
                )
            DB_PATH.parent.mkdir(parents=True, exist_ok=True)
            g.db = sqlite3.connect(DB_PATH)
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_: Any) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = get_db()
    execute_script(db, active_schema_sql())
    ensure_support_records_schema(db)
    db.commit()


def ensure_support_records_schema(db: Any) -> None:
    if USE_POSTGRES:
        execute_query(
            db,
            """
            ALTER TABLE support_records
            ADD COLUMN IF NOT EXISTS requester_name TEXT NOT NULL DEFAULT ''
            """,
        )
    else:
        columns = {
            row["name"]
            for row in execute_query(db, "PRAGMA table_info(support_records)").fetchall()
        }
        if "requester_name" not in columns:
            execute_query(
                db,
                """
                ALTER TABLE support_records
                ADD COLUMN requester_name TEXT NOT NULL DEFAULT ''
                """,
            )

    execute_query(
        db,
        """
        CREATE INDEX IF NOT EXISTS idx_support_requester
        ON support_records(requester_name)
        """,
    )


def ensure_default_services() -> None:
    db = get_db()
    existing_names = {
        row["name"]
        for row in execute_query(db, "SELECT name FROM services").fetchall()
    }
    missing_services = [
        service_name for service_name in DEFAULT_SERVICES if service_name not in existing_names
    ]
    if not missing_services:
        return

    timestamp = iso_now()
    execute_many(
        db,
        """
        INSERT INTO services (name, description, is_active, created_at, updated_at)
        VALUES (?, ?, 1, ?, ?)
        """,
        [
            (
                service_name,
                f"Theo dõi lượt hỗ trợ cho dịch vụ {service_name}.",
                timestamp,
                timestamp,
            )
            for service_name in missing_services
        ],
    )
    db.commit()


def seed_sample_data() -> None:
    db = get_db()
    ensure_default_services()

    entry_count = execute_query(
        db,
        "SELECT COUNT(*) AS total FROM support_records",
    ).fetchone()["total"]
    if entry_count > 0:
        return

    service_lookup = {
        row["name"]: row["id"]
        for row in execute_query(db, "SELECT id, name FROM services").fetchall()
    }
    # Seed trải đều sang tuần trước và tuần hiện tại để bộ lọc mặc định có số liệu ngay.
    base_date = current_week_range(date.today())[0]
    timestamp = iso_now()

    sample_entries = [
        (6, "Trường THCS Lê Quý Đôn", "Kế toán", "VNPT CA", "Hỗ trợ ký số", "", "Hằng", "Đã xử lý", ""),
        (6, "Công ty Thực phẩm An Khang", "Lãnh đạo", "VNPT CA", "Thiết lập thêm dải số hóa đơn cho chi nhánh.", "", "Duyên", "Đã xử lý", ""),
    ]

    execute_many(
        db,
        """
        INSERT INTO support_records (
            support_date, customer_name, requester_name, service_id, support_content, channel,
            supporter_name, status, notes, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                (base_date + timedelta(days=offset)).isoformat(),
                customer_name,
                requester_name,
                service_lookup[service_name],
                content,
                channel,
                supporter_name,
                status,
                notes,
                timestamp,
                timestamp,
            )
            for (
                offset,
                customer_name,
                requester_name,
                service_name,
                content,
                channel,
                supporter_name,
                status,
                notes,
            ) in sample_entries
        ],
    )
    db.commit()


def ensure_database() -> None:
    with app.app_context():
        init_db()
        seed_sample_data()


def serialize_service(row: Any) -> dict[str, Any]:
    return {
        "id": row["id"],
        "name": row["name"],
        "description": row["description"],
        "is_active": bool(row["is_active"]),
        "usage_count": row["usage_count"],
    }


def fetch_services(include_inactive: bool = True) -> list[dict[str, Any]]:
    db = get_db()
    sql = """
        SELECT
            services.id,
            services.name,
            services.description,
            services.is_active,
            COUNT(support_records.id) AS usage_count
        FROM services
        LEFT JOIN support_records ON support_records.service_id = services.id
    """
    params: list[Any] = []
    if not include_inactive:
        sql += " WHERE services.is_active = 1"
    sql += """
        GROUP BY services.id, services.name, services.description, services.is_active
        ORDER BY services.is_active DESC, LOWER(services.name) ASC
    """
    return [serialize_service(row) for row in execute_query(db, sql, params).fetchall()]


def fetch_service(service_id: int):
    db = get_db()
    return execute_query(
        db,
        """
        SELECT id, name, description, is_active
        FROM services
        WHERE id = ?
        """,
        (service_id,),
    ).fetchone()


def parse_filters(source: Any, default_to_current_week: bool = False) -> dict[str, Any]:
    date_from = normalize_text(source.get("date_from"))
    date_to = normalize_text(source.get("date_to"))
    service_id = normalize_text(source.get("service_id"))
    status = normalize_text(source.get("status"))
    supporter_name = normalize_text(source.get("supporter_name"))

    if default_to_current_week and not date_from and not date_to:
        current_from, current_to = current_week_range()
        date_from = current_from.isoformat()
        date_to = current_to.isoformat()

    if date_from:
        date.fromisoformat(date_from)
    if date_to:
        date.fromisoformat(date_to)
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    parsed_service_id = int(service_id) if service_id else None
    return {
        "date_from": date_from or None,
        "date_to": date_to or None,
        "service_id": parsed_service_id,
        "status": status or None,
        "supporter_name": supporter_name or None,
    }


def build_entry_query(filters: dict[str, Any]) -> tuple[str, list[Any]]:
    sql = """
        SELECT
            support_records.id,
            support_records.support_date,
            support_records.customer_name,
            support_records.requester_name,
            support_records.service_id,
            services.name AS service_name,
            support_records.support_content,
            support_records.channel,
            support_records.supporter_name,
            support_records.status,
            support_records.notes,
            support_records.created_at,
            support_records.updated_at
        FROM support_records
        JOIN services ON services.id = support_records.service_id
        WHERE 1 = 1
    """
    params: list[Any] = []
    if filters["date_from"]:
        sql += " AND support_records.support_date >= ?"
        params.append(filters["date_from"])
    if filters["date_to"]:
        sql += " AND support_records.support_date <= ?"
        params.append(filters["date_to"])
    if filters["service_id"]:
        sql += " AND support_records.service_id = ?"
        params.append(filters["service_id"])
    if filters["status"]:
        sql += " AND support_records.status = ?"
        params.append(filters["status"])
    if filters["supporter_name"]:
        sql += " AND support_records.supporter_name = ?"
        params.append(filters["supporter_name"])

    sql += " ORDER BY support_records.support_date DESC, support_records.id DESC"
    return sql, params


def fetch_entries(filters: dict[str, Any]) -> list[dict[str, Any]]:
    db = get_db()
    sql, params = build_entry_query(filters)
    rows = execute_query(db, sql, params).fetchall()
    entries = [dict(row) for row in rows]
    for entry in entries:
        entry["channel"] = canonical_channel_name(entry.get("channel"))
    return entries


def fetch_entry(entry_id: int) -> dict[str, Any] | None:
    db = get_db()
    row = execute_query(
        db,
        """
        SELECT
            support_records.id,
            support_records.support_date,
            support_records.customer_name,
            support_records.requester_name,
            support_records.service_id,
            services.name AS service_name,
            support_records.support_content,
            support_records.channel,
            support_records.supporter_name,
            support_records.status,
            support_records.notes,
            support_records.created_at,
            support_records.updated_at
        FROM support_records
        JOIN services ON services.id = support_records.service_id
        WHERE support_records.id = ?
        """,
        (entry_id,),
    ).fetchone()
    if not row:
        return None
    entry = dict(row)
    entry["channel"] = canonical_channel_name(entry.get("channel"))
    return entry


def fetch_supporters() -> list[str]:
    db = get_db()
    rows = execute_query(
        db,
        """
        SELECT supporter_name
        FROM support_records
        WHERE TRIM(supporter_name) <> ''
        GROUP BY supporter_name
        ORDER BY LOWER(supporter_name) ASC
        """
    ).fetchall()
    supporter_names = {
        supporter["name"]
        for supporter in SUPPORTER_OPTIONS
        if normalize_text(supporter.get("name"))
    }
    supporter_names.update(
        row["supporter_name"]
        for row in rows
        if normalize_text(row["supporter_name"])
    )
    return sorted(supporter_names, key=lambda value: value.lower())


def fetch_customers() -> list[str]:
    db = get_db()
    rows = execute_query(
        db,
        """
        SELECT customer_name
        FROM support_records
        WHERE TRIM(customer_name) <> ''
        GROUP BY customer_name
        ORDER BY LOWER(customer_name) ASC
        """
    ).fetchall()
    return [row["customer_name"] for row in rows]


def build_requester_summary(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    total_entries = len(entries)
    requester_counter = Counter(
        normalize_text(entry.get("requester_name", "")) or "Để trống"
        for entry in entries
    )
    rows: list[dict[str, Any]] = []
    handled_requesters: set[str] = set()

    for requester_name in REQUESTER_OPTION_NAMES:
        count = requester_counter.get(requester_name, 0)
        if not count:
            continue
        rows.append(
            {
                "requester_name": requester_name,
                "total_supports": count,
                "share": round((count / total_entries) * 100, 1) if total_entries else 0,
            }
        )
        handled_requesters.add(requester_name)

    for requester_name, count in sorted(
        requester_counter.items(),
        key=lambda item: (-item[1], item[0].lower()),
    ):
        if requester_name in handled_requesters:
            continue
        rows.append(
            {
                "requester_name": requester_name,
                "total_supports": count,
                "share": round((count / total_entries) * 100, 1) if total_entries else 0,
            }
        )

    return rows


def get_service_content_suggestions(service_name: str) -> list[str]:
    normalized_service_name = normalize_lookup_text(service_name)
    if not normalized_service_name:
        return []

    for suggestion_group in SERVICE_CONTENT_SUGGESTION_GROUPS:
        aliases = suggestion_group["aliases"]
        if any(
            normalize_lookup_text(alias) in normalized_service_name
            for alias in aliases
        ):
            return suggestion_group["suggestions"]

    return [
        f"Hỗ trợ đăng nhập và sử dụng chức năng cơ bản của dịch vụ {service_name}.",
        f"Hỗ trợ cập nhật cấu hình, plugin hoặc phần mềm liên quan đến dịch vụ {service_name}.",
        f"Xử lý lỗi phát sinh trong quá trình thao tác dịch vụ {service_name}.",
    ]


def build_content_suggestions_by_service(
    services: list[dict[str, Any]],
) -> dict[str, list[str]]:
    return {
        str(service["id"]): get_service_content_suggestions(service["name"])
        for service in services
    }


def build_dashboard(entries: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(entries)
    service_counter = Counter(entry["service_name"] for entry in entries)
    day_counter = Counter(entry["support_date"] for entry in entries)
    status_counter = Counter(entry["status"] for entry in entries)
    customer_counter = Counter(
        entry["customer_name"]
        for entry in entries
        if normalize_text(entry["customer_name"])
    )
    channel_counter = Counter(
        canonical_channel_name(entry["channel"])
        for entry in entries
        if canonical_channel_name(entry["channel"])
    )
    unresolved = (
        status_counter.get(STATUS_OPTIONS[1], 0)
        + status_counter.get(STATUS_OPTIONS[2], 0)
    )

    by_service = [
        {
            "name": name,
            "count": count,
            "share": round((count / total) * 100, 1) if total else 0,
        }
        for name, count in sorted(service_counter.items(), key=lambda item: (-item[1], item[0].lower()))
    ]
    by_day = [
        {
            "date": support_date,
            "label": format_display_date(support_date),
            "count": count,
            "weekday": date.fromisoformat(support_date).strftime("%A"),
        }
        for support_date, count in sorted(day_counter.items())
    ]
    by_status = [
        {
            "name": status,
            "count": status_counter.get(status, 0),
            "share": round((status_counter.get(status, 0) / total) * 100, 1) if total else 0,
        }
        for status in STATUS_OPTIONS
    ]
    by_channel = [
        {
            "name": channel,
            "count": channel_counter.get(channel, 0),
        }
        for channel in CHANNEL_OPTIONS
        if channel_counter.get(channel, 0)
    ]
    top_customers = [
        {
            "name": name,
            "count": count,
        }
        for name, count in sorted(customer_counter.items(), key=lambda item: (-item[1], item[0].lower()))[:5]
    ]

    return {
        "total_supports": total,
        "unique_customers": len(customer_counter),
        "services_covered": len(service_counter),
        "open_items": unresolved,
        "by_service": by_service,
        "by_day": by_day,
        "by_status": by_status,
        "by_channel": by_channel,
        "top_customers": top_customers,
    }


def build_service_summary(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in entries:
        grouped[entry["service_name"]].append(entry)

    rows: list[dict[str, Any]] = []
    for service_name, service_entries in sorted(
        grouped.items(),
        key=lambda item: (-len(item[1]), item[0].lower()),
    ):
        status_counter = Counter(entry["status"] for entry in service_entries)
        channel_counter = Counter(
            canonical_channel_name(entry["channel"])
            for entry in service_entries
            if canonical_channel_name(entry["channel"])
        )
        customer_counter = Counter(
            entry["customer_name"]
            for entry in service_entries
            if normalize_text(entry["customer_name"])
        )
        channel_counts = {
            channel: channel_counter.get(channel, 0)
            for channel in CHANNEL_OPTIONS
        }
        topic_samples: list[str] = []
        for entry in service_entries:
            topic = short_text(entry["support_content"], limit=75)
            if topic and topic not in topic_samples:
                topic_samples.append(topic)
            if len(topic_samples) == 3:
                break

        top_customer = ""
        if customer_counter:
            customer_name, customer_count = customer_counter.most_common(1)[0]
            top_customer = f"{customer_name} ({customer_count} lượt)"

        rows.append(
            {
                "service_name": service_name,
                "total_supports": len(service_entries),
                "unique_customers": len(customer_counter),
                "done_count": status_counter.get(STATUS_OPTIONS[0], 0),
                "in_progress_count": status_counter.get(STATUS_OPTIONS[1], 0),
                "waiting_count": status_counter.get(STATUS_OPTIONS[2], 0),
                "channel_counts": channel_counts,
                "top_customer": top_customer,
                "main_topics": "; ".join(topic_samples),
                "channel_summary": format_counter(channel_counter, CHANNEL_OPTIONS),
                "status_summary": format_counter(status_counter, STATUS_OPTIONS),
            }
        )
    return rows


def build_report_preview(entries: list[dict[str, Any]], filters: dict[str, Any]) -> dict[str, Any]:
    if not filters["date_from"] and entries:
        date_from = min(entry["support_date"] for entry in entries)
    else:
        date_from = filters["date_from"]

    if not filters["date_to"] and entries:
        date_to = max(entry["support_date"] for entry in entries)
    else:
        date_to = filters["date_to"]

    period_label = "toàn bộ dữ liệu"
    if date_from and date_to:
        period_label = f"từ ngày {format_display_date(date_from)} đến {format_display_date(date_to)}"
    elif date_from:
        period_label = f"từ ngày {format_display_date(date_from)}"
    elif date_to:
        period_label = f"đến ngày {format_display_date(date_to)}"

    if not entries:
        empty_text = f"Không có lượt hỗ trợ phát sinh trong khoảng {period_label}."
        return {
            "period_label": period_label,
            "intro": empty_text,
            "service_rows": [],
            "full_text": empty_text,
        }

    summary_rows = build_service_summary(entries)
    service_counter = Counter(entry["service_name"] for entry in entries)
    customer_counter = Counter(
        entry["customer_name"]
        for entry in entries
        if normalize_text(entry["customer_name"])
    )
    top_services = ", ".join(
        f"{name} ({count} lượt)"
        for name, count in service_counter.most_common(3)
    )
    intro = (
        f"Trong khoảng {period_label}, đã ghi nhận {len(entries)} lượt hỗ trợ cho "
        f"{len(customer_counter)} khách hàng/đơn vị, phát sinh trên {len(service_counter)} dịch vụ. "
        f"Các dịch vụ phát sinh nhiều nhất gồm: {top_services}."
    )

    bullet_lines: list[str] = []
    service_rows: list[dict[str, Any]] = []
    for row in summary_rows:
        suggestion = (
            f"Hỗ trợ dịch vụ {row['service_name']}: {row['total_supports']} lượt cho "
            f"{row['unique_customers']} khách hàng, kênh hỗ trợ gồm {row['channel_summary']}. "
            f"Nội dung chính: {row['main_topics'] or 'tiếp nhận và xử lý yêu cầu thường xuyên'}."
        )
        if row["in_progress_count"] or row["waiting_count"]:
            pending_total = row["in_progress_count"] + row["waiting_count"]
            suggestion += f" Còn {pending_total} lượt cần tiếp tục theo dõi."

        bullet_lines.append(f"- {suggestion}")
        service_rows.append(
            {
                "service_name": row["service_name"],
                "total_supports": row["total_supports"],
                "suggestion": suggestion,
                "status_summary": row["status_summary"],
                "channel_summary": row["channel_summary"],
            }
        )

    closing = "Có thể dùng các dòng trên để điền nhanh vào cột \"Nội dung cụ thể\" trong báo cáo tuần."
    full_text = "\n".join([intro, *bullet_lines, closing])

    return {
        "period_label": period_label,
        "intro": intro,
        "service_rows": service_rows,
        "full_text": full_text,
    }


def validate_service_payload(payload: dict[str, Any]) -> tuple[dict[str, Any], dict[str, str]]:
    errors: dict[str, str] = {}
    name = normalize_text(payload.get("name"))
    description = " ".join(str(payload.get("description") or "").strip().split())
    is_active_value = payload.get("is_active", True)
    is_active = bool(is_active_value)

    if not name:
        errors["name"] = "Tên dịch vụ không được để trống."

    return {
        "name": name,
        "description": description,
        "is_active": is_active,
    }, errors


def validate_entry_payload(payload: dict[str, Any]) -> tuple[dict[str, Any], dict[str, str]]:
    errors: dict[str, str] = {}
    support_date = normalize_text(payload.get("support_date"))
    customer_name = " ".join(str(payload.get("customer_name") or "").strip().split())
    requester_name = normalize_text(payload.get("requester_name"))
    service_id_raw = str(payload.get("service_id") or "").strip()
    support_content = str(payload.get("support_content") or "").strip()
    channel = canonical_channel_name(payload.get("channel"))
    supporter_name = " ".join(str(payload.get("supporter_name") or "").strip().split())
    status = normalize_text(payload.get("status"))
    notes = str(payload.get("notes") or "").strip()

    if not support_date:
        errors["support_date"] = "Ngày hỗ trợ là bắt buộc."
    else:
        try:
            date.fromisoformat(support_date)
        except ValueError:
            errors["support_date"] = "Ngày hỗ trợ không đúng định dạng."

    if not service_id_raw:
        errors["service_id"] = "Vui lòng chọn dịch vụ."
        service_id = None
    else:
        try:
            service_id = int(service_id_raw)
        except ValueError:
            service_id = None
            errors["service_id"] = "Dịch vụ không hợp lệ."

    if service_id is not None and fetch_service(service_id) is None:
        errors["service_id"] = "Dịch vụ không tồn tại."


    if channel not in CHANNEL_OPTIONS:
        errors["channel"] = "Kênh hỗ trợ không hợp lệ."

    if not supporter_name:
        errors["supporter_name"] = "Người hỗ trợ là bắt buộc."

    if status not in STATUS_OPTIONS:
        errors["status"] = "Trạng thái không hợp lệ."

    return {
        "support_date": support_date,
        "customer_name": customer_name,
        "requester_name": requester_name,
        "service_id": service_id,
        "support_content": support_content,
        "channel": channel,
        "supporter_name": supporter_name,
        "status": status,
        "notes": notes,
    }, errors


def excel_is_available() -> bool:
    return Workbook is not None


def build_excel_report(
    entries: list[dict[str, Any]],
    filters: dict[str, Any],
    report_preview: dict[str, Any],
) -> io.BytesIO:
    if not excel_is_available():
        raise RuntimeError("Thiếu thư viện openpyxl để xuất Excel.")

    summary_rows = build_service_summary(entries)
    requester_rows = build_requester_summary(entries)
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1F6F5F")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 1: dữ liệu chi tiết để rà soát từng lượt hỗ trợ.
    detail_sheet = workbook.active
    detail_sheet.title = "Chi tiet ho tro"
    detail_sheet["A1"] = APP_TITLE
    detail_sheet["A1"].font = title_font
    detail_sheet["A2"] = f"Khoảng báo cáo: {report_preview['period_label']}"
    detail_sheet["A3"] = f"Tổng lượt hỗ trợ: {len(entries)}"

    detail_headers = [
        "Ngày hỗ trợ",
        "Khách hàng / Đơn vị",
        "Người gửi hỗ trợ",
        "Dịch vụ hỗ trợ",
        "Nội dung hỗ trợ",
        "Kênh hỗ trợ",
        "Người hỗ trợ",
        "Trạng thái",
        "Ghi chú",
    ]
    detail_sheet.append([])
    detail_sheet.append(detail_headers)
    for column in "ABCDEFGHI":
        detail_sheet[f"{column}5"].fill = header_fill
        detail_sheet[f"{column}5"].font = header_font

    for entry in sorted(entries, key=lambda item: (item["support_date"], item["id"])):
        detail_sheet.append(
            [
                format_display_date(entry["support_date"]),
                entry["customer_name"],
                entry.get("requester_name", ""),
                entry["service_name"],
                entry["support_content"],
                entry["channel"],
                entry["supporter_name"],
                entry["status"],
                entry["notes"],
            ]
        )

    for row in detail_sheet.iter_rows(min_row=6, max_col=9):
        for cell in row:
            cell.alignment = wrap_alignment
    detail_sheet.freeze_panes = "A6"
    detail_sheet.column_dimensions["A"].width = 14
    detail_sheet.column_dimensions["B"].width = 28
    detail_sheet.column_dimensions["C"].width = 20
    detail_sheet.column_dimensions["D"].width = 22
    detail_sheet.column_dimensions["E"].width = 48
    detail_sheet.column_dimensions["F"].width = 18
    detail_sheet.column_dimensions["G"].width = 20
    detail_sheet.column_dimensions["H"].width = 16
    detail_sheet.column_dimensions["I"].width = 30

    summary_sheet = workbook.create_sheet("Tong hop dich vu")
    summary_sheet["A1"] = "Tổng hợp số lượt hỗ trợ theo dịch vụ"
    summary_sheet["A1"].font = title_font
    summary_sheet["A2"] = f"Khoảng báo cáo: {report_preview['period_label']}"
    summary_headers = [
        "Dịch vụ",
        "Tổng lượt",
        "Số khách hàng",
        "Đã xử lý",
        "Đang xử lý",
        "Chờ phản hồi",
        *CHANNEL_OPTIONS,
        "Khách hàng nổi bật",
        "Nội dung chính",
    ]
    summary_sheet.append([])
    summary_sheet.append(summary_headers)
    for column in "ABCDEFGHIJKLM":
        summary_sheet[f"{column}4"].fill = header_fill
        summary_sheet[f"{column}4"].font = header_font

    for row in summary_rows:
        summary_sheet.append(
            [
                row["service_name"],
                row["total_supports"],
                row["unique_customers"],
                row["done_count"],
                row["in_progress_count"],
                row["waiting_count"],
                *[row["channel_counts"].get(channel, 0) for channel in CHANNEL_OPTIONS],
                row["top_customer"],
                row["main_topics"],
            ]
        )
    for row in summary_sheet.iter_rows(min_row=5, max_col=13):
        for cell in row:
            cell.alignment = wrap_alignment
    summary_sheet.freeze_panes = "A5"
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 12
    summary_sheet.column_dimensions["C"].width = 14
    summary_sheet.column_dimensions["D"].width = 12
    summary_sheet.column_dimensions["E"].width = 14
    summary_sheet.column_dimensions["F"].width = 14
    summary_sheet.column_dimensions["G"].width = 12
    summary_sheet.column_dimensions["H"].width = 12
    summary_sheet.column_dimensions["I"].width = 10
    summary_sheet.column_dimensions["J"].width = 12
    summary_sheet.column_dimensions["K"].width = 8
    summary_sheet.column_dimensions["L"].width = 28
    summary_sheet.column_dimensions["M"].width = 48

    requester_sheet = workbook.create_sheet("Thong ke nguoi gui")
    requester_sheet["A1"] = "Thống kê số lượng theo người gửi hỗ trợ"
    requester_sheet["A1"].font = title_font
    requester_sheet["A2"] = f"Khoảng báo cáo: {report_preview['period_label']}"
    requester_headers = [
        "Người gửi hỗ trợ",
        "Số lượt hỗ trợ",
        "Tỷ lệ (%)",
    ]
    requester_sheet.append([])
    requester_sheet.append(requester_headers)
    for column in "ABC":
        requester_sheet[f"{column}4"].fill = header_fill
        requester_sheet[f"{column}4"].font = header_font

    for row in requester_rows:
        requester_sheet.append(
            [
                row["requester_name"],
                row["total_supports"],
                row["share"],
            ]
        )

    for row in requester_sheet.iter_rows(min_row=5, max_col=3):
        for cell in row:
            cell.alignment = wrap_alignment
    requester_sheet.freeze_panes = "A5"
    requester_sheet.column_dimensions["A"].width = 28
    requester_sheet.column_dimensions["B"].width = 16
    requester_sheet.column_dimensions["C"].width = 12

    narrative_sheet = workbook.create_sheet("Noi dung goi y")
    narrative_sheet["A1"] = "Nội dung gợi ý để copy vào báo cáo tuần"
    narrative_sheet["A1"].font = title_font
    narrative_sheet["A2"] = f"Khoảng báo cáo: {report_preview['period_label']}"
    narrative_sheet["A4"] = "Đoạn tổng hợp nhanh"
    narrative_sheet["A4"].font = Font(bold=True)
    narrative_sheet["A5"] = report_preview["full_text"]
    narrative_sheet["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    narrative_sheet.merge_cells("A5:G10")

    narrative_headers = [
        "Dịch vụ",
        "Tổng lượt",
        "Kênh hỗ trợ",
        "Trạng thái",
        "Nội dung gợi ý",
    ]
    narrative_sheet["A12"] = "Dòng gợi ý theo từng dịch vụ"
    narrative_sheet["A12"].font = Font(bold=True)
    narrative_sheet.append([])
    narrative_sheet.append(narrative_headers)
    for column in "ABCDE":
        narrative_sheet[f"{column}14"].fill = header_fill
        narrative_sheet[f"{column}14"].font = header_font

    for row in report_preview["service_rows"]:
        narrative_sheet.append(
            [
                row["service_name"],
                row["total_supports"],
                row["channel_summary"],
                row["status_summary"],
                row["suggestion"],
            ]
        )

    for row in narrative_sheet.iter_rows(min_row=15, max_col=5):
        for cell in row:
            cell.alignment = wrap_alignment
    narrative_sheet.column_dimensions["A"].width = 24
    narrative_sheet.column_dimensions["B"].width = 12
    narrative_sheet.column_dimensions["C"].width = 26
    narrative_sheet.column_dimensions["D"].width = 24
    narrative_sheet.column_dimensions["E"].width = 70
    narrative_sheet.freeze_panes = "A15"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def make_csv_content(headers: list[str], rows: list[list[Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue()


def build_csv_bundle(
    entries: list[dict[str, Any]],
    report_preview: dict[str, Any],
) -> io.BytesIO:
    summary_rows = build_service_summary(entries)
    requester_rows = build_requester_summary(entries)
    detail_csv = make_csv_content(
        [
            "Ngày hỗ trợ",
            "Khách hàng / Đơn vị",
            "Người gửi hỗ trợ",
            "Dịch vụ hỗ trợ",
            "Nội dung hỗ trợ",
            "Kênh hỗ trợ",
            "Người hỗ trợ",
            "Trạng thái",
            "Ghi chú",
        ],
        [
            [
                format_display_date(entry["support_date"]),
                entry["customer_name"],
                entry.get("requester_name", ""),
                entry["service_name"],
                entry["support_content"],
                entry["channel"],
                entry["supporter_name"],
                entry["status"],
                entry["notes"],
            ]
            for entry in sorted(entries, key=lambda item: (item["support_date"], item["id"]))
        ],
    )

    requester_csv = make_csv_content(
        ["Người gửi hỗ trợ", "Số lượt hỗ trợ", "Tỷ lệ (%)"],
        [
            [
                row["requester_name"],
                row["total_supports"],
                row["share"],
            ]
            for row in requester_rows
        ],
    )

    summary_csv = make_csv_content(
        [
            "Dịch vụ",
            "Tổng lượt",
            "Số khách hàng",
            "Đã xử lý",
            "Đang xử lý",
            "Chờ phản hồi",
            *CHANNEL_OPTIONS,
            "Khách hàng nổi bật",
            "Nội dung chính",
        ],
        [
            [
                row["service_name"],
                row["total_supports"],
                row["unique_customers"],
                row["done_count"],
                row["in_progress_count"],
                row["waiting_count"],
                *[row["channel_counts"].get(channel, 0) for channel in CHANNEL_OPTIONS],
                row["top_customer"],
                row["main_topics"],
            ]
            for row in summary_rows
        ],
    )

    suggestion_csv = make_csv_content(
        ["Dịch vụ", "Tổng lượt", "Kênh hỗ trợ", "Trạng thái", "Nội dung gợi ý"],
        [
            [
                row["service_name"],
                row["total_supports"],
                row["channel_summary"],
                row["status_summary"],
                row["suggestion"],
            ]
            for row in report_preview["service_rows"]
        ],
    )

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("01_chi_tiet_ho_tro.csv", detail_csv.encode("utf-8-sig"))
        archive.writestr("02_tong_hop_dich_vu.csv", summary_csv.encode("utf-8-sig"))
        archive.writestr("03_thong_ke_nguoi_gui.csv", requester_csv.encode("utf-8-sig"))
        archive.writestr("04_noi_dung_goi_y.csv", suggestion_csv.encode("utf-8-sig"))
        archive.writestr("04_noi_dung_goi_y.txt", report_preview["full_text"].encode("utf-8-sig"))
    output.seek(0)
    return output


def report_filename(prefix: str, filters: dict[str, Any], extension: str) -> str:
    date_from = filters["date_from"] or "tu_dau"
    date_to = filters["date_to"] or "den_cuoi"
    return f"{prefix}_{date_from}_{date_to}.{extension}"


@app.route("/")
def index() -> str:
    week_start, week_end = current_week_range()
    return render_template(
        "index.html",
        app_title=APP_TITLE,
        channel_options=CHANNEL_OPTIONS,
        requester_options=REQUESTER_OPTIONS,
        supporter_options=SUPPORTER_OPTIONS,
        status_options=STATUS_OPTIONS,
        default_filters={
            "date_from": week_start.isoformat(),
            "date_to": week_end.isoformat(),
        },
    )


@app.get("/api/bootstrap")
def api_bootstrap():
    filters = parse_filters(request.args, default_to_current_week=True)
    entries = fetch_entries(filters)
    services = fetch_services(include_inactive=True)
    return jsonify(
        {
            "filters": filters,
            "services": services,
            "content_suggestions": build_content_suggestions_by_service(services),
            "supporters": fetch_supporters(),
            "customers": fetch_customers(),
            "entries": entries,
            "dashboard": build_dashboard(entries),
            "report_preview": build_report_preview(entries, filters),
            "channels": CHANNEL_OPTIONS,
            "requester_options": REQUESTER_OPTIONS,
            "statuses": STATUS_OPTIONS,
        }
    )


@app.get("/api/services")
def api_services():
    include_inactive = request.args.get("include_inactive", "1") != "0"
    return jsonify(fetch_services(include_inactive=include_inactive))


@app.post("/api/services")
def create_service():
    payload = request.get_json(silent=True) or {}
    cleaned, errors = validate_service_payload(payload)
    if errors:
        return jsonify({"errors": errors}), 400

    db = get_db()
    timestamp = iso_now()
    try:
        insert_sql = """
            INSERT INTO services (name, description, is_active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
        """
        if USE_POSTGRES:
            insert_sql += " RETURNING id"
        cursor = execute_query(
            db,
            insert_sql,
            (
                cleaned["name"],
                cleaned["description"],
                1 if cleaned["is_active"] else 0,
                timestamp,
                timestamp,
            ),
        )
        service_id = cursor.fetchone()["id"] if USE_POSTGRES else cursor.lastrowid
        db.commit()
    except DB_INTEGRITY_ERRORS:
        return jsonify({"errors": {"name": "Dịch vụ đã tồn tại."}}), 400

    row = fetch_service(service_id)
    return jsonify({"message": "Đã tạo dịch vụ.", "service": dict(row) if row else None}), 201


@app.put("/api/services/<int:service_id>")
def update_service(service_id: int):
    if fetch_service(service_id) is None:
        abort(404)

    payload = request.get_json(silent=True) or {}
    cleaned, errors = validate_service_payload(payload)
    if errors:
        return jsonify({"errors": errors}), 400

    db = get_db()
    try:
        execute_query(
            db,
            """
            UPDATE services
            SET name = ?, description = ?, is_active = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                cleaned["name"],
                cleaned["description"],
                1 if cleaned["is_active"] else 0,
                iso_now(),
                service_id,
            ),
        )
        db.commit()
    except DB_INTEGRITY_ERRORS:
        return jsonify({"errors": {"name": "Tên dịch vụ đã tồn tại."}}), 400

    return jsonify({"message": "Đã cập nhật dịch vụ."})


@app.delete("/api/services/<int:service_id>")
def delete_service(service_id: int):
    service = fetch_service(service_id)
    if service is None:
        abort(404)

    db = get_db()
    usage_count = execute_query(
        db,
        "SELECT COUNT(*) AS total FROM support_records WHERE service_id = ?",
        (service_id,),
    ).fetchone()["total"]

    if usage_count > 0:
        execute_query(
            db,
            """
            UPDATE services
            SET is_active = 0, updated_at = ?
            WHERE id = ?
            """,
            (iso_now(), service_id),
        )
        db.commit()
        return jsonify(
            {
                "message": "Dịch vụ đã được ngừng sử dụng vì đang có dữ liệu lịch sử liên quan."
            }
        )

    execute_query(db, "DELETE FROM services WHERE id = ?", (service_id,))
    db.commit()
    return jsonify({"message": "Đã xóa dịch vụ."})


@app.post("/api/entries")
def create_entry():
    payload = request.get_json(silent=True) or {}
    cleaned, errors = validate_entry_payload(payload)
    if errors:
        return jsonify({"errors": errors}), 400

    db = get_db()
    timestamp = iso_now()
    insert_sql = """
        INSERT INTO support_records (
            support_date, customer_name, requester_name, service_id, support_content, channel,
            supporter_name, status, notes, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    if USE_POSTGRES:
        insert_sql += " RETURNING id"
    cursor = execute_query(
        db,
        insert_sql,
        (
            cleaned["support_date"],
            cleaned["customer_name"],
            cleaned["requester_name"],
            cleaned["service_id"],
            cleaned["support_content"],
            cleaned["channel"],
            cleaned["supporter_name"],
            cleaned["status"],
            cleaned["notes"],
            timestamp,
            timestamp,
        ),
    )
    entry_id = cursor.fetchone()["id"] if USE_POSTGRES else cursor.lastrowid
    db.commit()
    created = fetch_entry(entry_id)
    return jsonify({"message": "Đã thêm lượt hỗ trợ.", "entry": created}), 201


@app.put("/api/entries/<int:entry_id>")
def update_entry(entry_id: int):
    if fetch_entry(entry_id) is None:
        abort(404)

    payload = request.get_json(silent=True) or {}
    cleaned, errors = validate_entry_payload(payload)
    if errors:
        return jsonify({"errors": errors}), 400

    db = get_db()
    execute_query(
        db,
        """
        UPDATE support_records
        SET support_date = ?, customer_name = ?, requester_name = ?, service_id = ?, support_content = ?,
            channel = ?, supporter_name = ?, status = ?, notes = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            cleaned["support_date"],
            cleaned["customer_name"],
            cleaned["requester_name"],
            cleaned["service_id"],
            cleaned["support_content"],
            cleaned["channel"],
            cleaned["supporter_name"],
            cleaned["status"],
            cleaned["notes"],
            iso_now(),
            entry_id,
        ),
    )
    db.commit()
    return jsonify({"message": "Đã cập nhật lượt hỗ trợ.", "entry": fetch_entry(entry_id)})


@app.delete("/api/entries/<int:entry_id>")
def delete_entry(entry_id: int):
    if fetch_entry(entry_id) is None:
        abort(404)

    db = get_db()
    execute_query(db, "DELETE FROM support_records WHERE id = ?", (entry_id,))
    db.commit()
    return jsonify({"message": "Đã xóa lượt hỗ trợ."})


@app.get("/api/report-preview")
def api_report_preview():
    filters = parse_filters(request.args, default_to_current_week=True)
    entries = fetch_entries(filters)
    return jsonify(build_report_preview(entries, filters))


@app.get("/export/report.xlsx")
def export_excel():
    filters = parse_filters(request.args, default_to_current_week=True)
    entries = fetch_entries(filters)
    preview = build_report_preview(entries, filters)
    try:
        output = build_excel_report(entries, filters, preview)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 500

    return send_file(
        output,
        as_attachment=True,
        download_name=report_filename("bao_cao_ho_tro", filters, "xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/export/report.csv")
def export_csv():
    filters = parse_filters(request.args, default_to_current_week=True)
    entries = fetch_entries(filters)
    preview = build_report_preview(entries, filters)
    output = build_csv_bundle(entries, preview)
    return send_file(
        output,
        as_attachment=True,
        download_name=report_filename("bao_cao_ho_tro", filters, "zip"),
        mimetype="application/zip",
    )


ensure_database()


if __name__ == "__main__":
    app.run(
        host=os.getenv("APP_HOST", "0.0.0.0"),
        port=int(os.getenv("APP_PORT", "5000")),
        debug=os.getenv("APP_DEBUG", "1") == "1",
    )
